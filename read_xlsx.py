import openpyxl

import os
# MSYS2/Git Bash path
path = "E:/AAA_Hermes_Shared_Workspace/temp_list.xlsx"
# Verify file exists
print(f"File exists: {os.path.exists(path)}")
print(f"File size: {os.path.getsize(path)}")
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
print(f'Sheet: {ws.title}')
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
print('---')
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    print(list(row))
